from flask import Blueprint, render_template, redirect, url_for, flash, request, current_app, jsonify
from flask_login import login_required, current_user
from app.models import db, Colaborador, SorteioSemanal, SorteioColaborador, Premio
from app.forms.manager import UploadColaboradoresForm, SorteioColaboradorForm, ColaboradorForm
from app.utils import get_brazil_datetime, get_brazil_date, format_brazil_datetime
from datetime import datetime, date, timedelta
from functools import wraps
import os
import random
import openpyxl
import json

manager_bp = Blueprint('manager', __name__)

def detectar_formato_planilha(sheet):
    """
    Detecta automaticamente o formato da planilha Excel baseado na estrutura das colunas
    
    Formato 1 (5 colunas): A=Unidade, B=Bandeira, C=Matrícula, D=Nome, E=Setor
    Formato 2 (6+ colunas): A=Filial, B=Unidade, C=Bandeira, D=Matrícula, E=Nome, F=Setor
    """
    try:
        # Analisa a primeira linha de dados (linha 2, já que linha 1 é cabeçalho)
        primeira_linha = list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        
        if not primeira_linha:
            # Fallback para formato 2 se não conseguir detectar
            return {
                "nome": "Formato 2 (Padrão) - 6 colunas",
                "col_matricula": 3,  # Coluna D
                "col_nome": 4,       # Coluna E
                "col_setor": 5,      # Coluna F
                "min_colunas": 6
            }
        
        # Verifica se a coluna A (índice 0) é numérica (indica formato 2)
        # Formato 2: A = número da filial (ex: 20202, 131329)
        col_a = primeira_linha[0] if len(primeira_linha) > 0 else None
        if col_a and str(col_a).strip().isdigit():
            return {
                "nome": "Formato 2 (6 colunas) - A:Filial, B:Unidade, C:Bandeira, D:Matrícula, E:Nome, F:Setor",
                "col_matricula": 3,  # Coluna D
                "col_nome": 4,       # Coluna E
                "col_setor": 5,      # Coluna F
                "min_colunas": 6
            }
        
        # Verifica se a coluna A contém código de loja (indica formato 1)
        # Formato 1: A = código da loja (ex: "BIG01 - 106 NORTE")
        if col_a and ("BIG" in str(col_a) or "ULTRA" in str(col_a)) and "-" in str(col_a):
            return {
                "nome": "Formato 1 (5 colunas) - A:Unidade, B:Bandeira, C:Matrícula, D:Nome, E:Setor",
                "col_matricula": 2,  # Coluna C
                "col_nome": 3,       # Coluna D
                "col_setor": 4,      # Coluna E
                "min_colunas": 5
            }
        
        # Se não conseguir detectar claramente, usa formato 2 como padrão
        return {
            "nome": "Formato 2 (Padrão detectado) - 6 colunas",
            "col_matricula": 3,  # Coluna D
            "col_nome": 4,       # Coluna E
            "col_setor": 5,      # Coluna F
            "min_colunas": 6
        }
        
    except Exception as e:
        # Em caso de erro, retorna formato 2 como fallback
        return {
            "nome": f"Formato 2 (Fallback após erro: {str(e)})",
            "col_matricula": 3,  # Coluna D
            "col_nome": 4,       # Coluna E
            "col_setor": 5,      # Coluna F
            "min_colunas": 6
        }

def manager_required(f):
    """Decorator para verificar se usuário é assistente"""
    @wraps(f)
    @login_required
    def decorated_function(*args, **kwargs):
        if current_user.tipo != 'assistente':
            flash('Acesso negado. Apenas assistentes de loja.', 'danger')
            return redirect(url_for('main.index'))
        return f(*args, **kwargs)
    return decorated_function

def allowed_file(filename):
    """Verifica se arquivo é permitido"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in current_app.config['ALLOWED_EXTENSIONS']

@manager_bp.route('/dashboard')
@manager_required
def dashboard():
    """Dashboard do assistente"""
    if not current_user.loja:
        flash('Você não está associado a nenhuma loja. Contate o administrador.', 'warning')
        return redirect(url_for('main.index'))
    
    # Verifica se a loja foi sorteada esta semana
    sorteio_atual = SorteioSemanal.query.order_by(SorteioSemanal.semana_inicio.desc()).first()
    loja_sorteada = False
    colaboradores_sorteados = []
    
    if sorteio_atual:
        # Verifica se a loja do assistente foi sorteada
        if (current_user.loja_id == sorteio_atual.loja_big_id or 
            current_user.loja_id == sorteio_atual.loja_ultra_id):
            loja_sorteada = True
            
            # Busca colaboradores já sorteados
            colaboradores_sorteados = SorteioColaborador.query.join(
                Colaborador
            ).filter(
                SorteioColaborador.sorteio_semanal_id == sorteio_atual.id,
                Colaborador.loja_id == current_user.loja_id
            ).all()
    
        # Estatísticas da loja
    total_colaboradores = Colaborador.query.filter_by(
        loja_id=current_user.loja_id,
        apto=True
    ).count()

    # Verifica prêmios disponíveis para sorteio
    premios_disponiveis_count = 0
    if sorteio_atual and loja_sorteada:
        # Prêmios já sorteados
        premios_ja_sorteados = db.session.query(SorteioColaborador.premio_id).join(Colaborador).filter(
            SorteioColaborador.sorteio_semanal_id == sorteio_atual.id,
            Colaborador.loja_id == current_user.loja_id
        ).subquery()

        # Conta prêmios disponíveis
        premios_disponiveis_count = Premio.query.filter(
            Premio.ativo == True,
            Premio.data_evento >= date.today(),
            db.or_(Premio.loja_id == current_user.loja_id, Premio.loja_id.is_(None)),
            ~Premio.id.in_(db.select(premios_ja_sorteados).scalar_subquery())
        ).count()

    return render_template('manager/dashboard.html',
                         loja=current_user.loja,
                         loja_sorteada=loja_sorteada,
                         sorteio_atual=sorteio_atual,
                         colaboradores_sorteados=colaboradores_sorteados,
                         total_colaboradores=total_colaboradores,
                         premios_disponiveis=premios_disponiveis_count)

@manager_bp.route('/colaboradores')
@manager_required
def colaboradores():
    """Lista colaboradores da loja com ordenação"""
    if not current_user.loja:
        flash('Você não está associado a nenhuma loja.', 'warning')
        return redirect(url_for('main.index'))
    
    # Parâmetros de ordenação
    sort_by = request.args.get('sort', 'nome')
    order = request.args.get('order', 'asc')
    
    # Monta query base
    query = Colaborador.query.filter_by(loja_id=current_user.loja_id)
    
    # Aplica ordenação
    if sort_by == 'matricula':
        if order == 'desc':
            query = query.order_by(Colaborador.matricula.desc())
        else:
            query = query.order_by(Colaborador.matricula.asc())
    elif sort_by == 'setor':
        if order == 'desc':
            query = query.order_by(Colaborador.setor.desc(), Colaborador.nome.asc())
        else:
            query = query.order_by(Colaborador.setor.asc(), Colaborador.nome.asc())
    else:  # nome (padrão)
        if order == 'desc':
            query = query.order_by(Colaborador.nome.desc())
        else:
            query = query.order_by(Colaborador.nome.asc())
    
    colaboradores = query.all()
    
    return render_template('manager/colaboradores.html', 
                         colaboradores=colaboradores,
                         loja=current_user.loja,
                         current_sort=sort_by,
                         current_order=order)

@manager_bp.route('/colaboradores/novo', methods=['GET', 'POST'])
@manager_required
def novo_colaborador():
    """Criar novo colaborador"""
    if not current_user.loja:
        flash('Você não está associado a nenhuma loja.', 'warning')
        return redirect(url_for('main.index'))
    
    form = ColaboradorForm()
    
    # Configura choices para campo loja_id (mesmo que não seja usado no template)
    form.loja_id.choices = [(current_user.loja_id, current_user.loja.nome)]
    
    if form.validate_on_submit():
        # Verifica se matrícula já existe na loja
        existente = Colaborador.query.filter_by(
            matricula=form.matricula.data,
            loja_id=current_user.loja_id
        ).first()
        
        if existente:
            flash('Já existe um colaborador com esta matrícula na sua loja!', 'danger')
            return render_template('manager/colaborador_form.html', form=form, titulo='Novo Colaborador')
        
        colaborador = Colaborador(
            matricula=form.matricula.data,
            nome=form.nome.data,
            setor=form.setor.data,
            loja_id=current_user.loja_id,
            apto=form.apto.data,
            ultima_atualizacao=get_brazil_datetime()
        )
        
        db.session.add(colaborador)
        db.session.commit()
        
        flash(f'Colaborador {colaborador.nome} criado com sucesso!', 'success')
        return redirect(url_for('manager.colaboradores'))
    
    return render_template('manager/colaborador_form.html', form=form, titulo='Novo Colaborador')

@manager_bp.route('/colaboradores/<int:id>/editar', methods=['GET', 'POST'])
@manager_required
def editar_colaborador(id):
    """Editar colaborador"""
    colaborador = Colaborador.query.filter_by(id=id, loja_id=current_user.loja_id).first_or_404()
    
    form = ColaboradorForm(obj=colaborador)
    
    # Configura choices para campo loja_id (mesmo que não seja usado no template)
    form.loja_id.choices = [(current_user.loja_id, current_user.loja.nome)]
    
    if form.validate_on_submit():
        # Verifica se matrícula já existe (exceto o próprio)
        existente = Colaborador.query.filter_by(
            matricula=form.matricula.data,
            loja_id=current_user.loja_id
        ).filter(Colaborador.id != id).first()
        
        if existente:
            flash('Já existe outro colaborador com esta matrícula na sua loja!', 'danger')
            return render_template('manager/colaborador_form.html', form=form, titulo='Editar Colaborador')
        
        colaborador.matricula = form.matricula.data
        colaborador.nome = form.nome.data
        colaborador.setor = form.setor.data
        colaborador.apto = form.apto.data
        colaborador.ultima_atualizacao = get_brazil_datetime()
        
        db.session.commit()
        
        flash(f'Colaborador {colaborador.nome} atualizado com sucesso!', 'success')
        return redirect(url_for('manager.colaboradores'))
    
    return render_template('manager/colaborador_form.html', form=form, titulo='Editar Colaborador')

@manager_bp.route('/colaboradores/<int:id>/toggle')
@manager_required
def toggle_colaborador(id):
    """Ativar/desativar colaborador"""
    colaborador = Colaborador.query.filter_by(id=id, loja_id=current_user.loja_id).first_or_404()
    
    colaborador.apto = not colaborador.apto
    colaborador.ultima_atualizacao = get_brazil_datetime()
    db.session.commit()
    
    status = 'habilitado' if colaborador.apto else 'desabilitado'
    flash(f'Colaborador {colaborador.nome} foi {status} para sorteios.', 'success')
    
    return redirect(url_for('manager.colaboradores'))

@manager_bp.route('/colaboradores/<int:id>/excluir')
@manager_required
def excluir_colaborador(id):
    """Excluir colaborador"""
    colaborador = Colaborador.query.filter_by(id=id, loja_id=current_user.loja_id).first_or_404()
    
    # Verifica se o colaborador tem sorteios
    tem_sorteios = SorteioColaborador.query.filter_by(colaborador_id=id).first()
    
    if tem_sorteios:
        flash('Não é possível excluir este colaborador pois ele já participou de sorteios.', 'warning')
        return redirect(url_for('manager.colaboradores'))
    
    nome = colaborador.nome
    db.session.delete(colaborador)
    db.session.commit()
    
    flash(f'Colaborador {nome} foi excluído com sucesso.', 'success')
    return redirect(url_for('manager.colaboradores'))

@manager_bp.route('/colaboradores/acao-lote', methods=['POST'])
@manager_required
def acao_lote():
    """Aplicar ação em lote para colaboradores selecionados"""
    if not current_user.loja:
        flash('Você não está associado a nenhuma loja.', 'warning')
        return redirect(url_for('main.index'))
    
    colaborador_ids = request.form.getlist('colaborador_ids')
    acao = request.form.get('acao')
    
    if not colaborador_ids:
        flash('Nenhum colaborador selecionado.', 'warning')
        return redirect(url_for('manager.colaboradores'))
    
    if not acao:
        flash('Nenhuma ação selecionada.', 'warning')
        return redirect(url_for('manager.colaboradores'))
    
    # Converte IDs para inteiros
    try:
        ids = [int(id) for id in colaborador_ids]
    except ValueError:
        flash('IDs de colaboradores inválidos.', 'danger')
        return redirect(url_for('manager.colaboradores'))
    
    # Busca colaboradores da loja
    colaboradores = Colaborador.query.filter(
        Colaborador.id.in_(ids),
        Colaborador.loja_id == current_user.loja_id
    ).all()
    
    if not colaboradores:
        flash('Colaboradores não encontrados.', 'warning')
        return redirect(url_for('manager.colaboradores'))
    
    sucesso = 0
    erros = 0
    
    if acao == 'ativar':
        for colaborador in colaboradores:
            colaborador.apto = True
            colaborador.ultima_atualizacao = get_brazil_datetime()
            sucesso += 1
        flash(f'{sucesso} colaboradores foram habilitados para sorteios.', 'success')
        
    elif acao == 'desativar':
        for colaborador in colaboradores:
            colaborador.apto = False
            colaborador.ultima_atualizacao = get_brazil_datetime()
            sucesso += 1
        flash(f'{sucesso} colaboradores foram desabilitados para sorteios.', 'success')
        
    elif acao == 'excluir':
        for colaborador in colaboradores:
            # Verifica se tem sorteios
            tem_sorteios = SorteioColaborador.query.filter_by(colaborador_id=colaborador.id).first()
            if not tem_sorteios:
                db.session.delete(colaborador)
                sucesso += 1
            else:
                erros += 1
        
        if sucesso > 0:
            flash(f'{sucesso} colaboradores foram excluídos.', 'success')
        if erros > 0:
            flash(f'{erros} colaboradores não puderam ser excluídos (possuem histórico de sorteios).', 'warning')
    
    else:
        flash('Ação inválida.', 'danger')
        return redirect(url_for('manager.colaboradores'))
    
    db.session.commit()
    return redirect(url_for('manager.colaboradores'))

@manager_bp.route('/colaboradores/upload', methods=['GET', 'POST'])
@manager_required
def upload_colaboradores():
    """Upload de planilha de colaboradores - SUBSTITUI TODOS"""
    if not current_user.loja:
        flash('Você não está associado a nenhuma loja.', 'warning')
        return redirect(url_for('main.index'))
    
    form = UploadColaboradoresForm()
    
    if form.validate_on_submit():
        arquivo = form.arquivo.data
        
        try:
            # Carrega o arquivo Excel
            workbook = openpyxl.load_workbook(arquivo)
            sheet = workbook.active
            
            # Detecta automaticamente o formato da planilha
            formato_detectado = detectar_formato_planilha(sheet)
            flash(f'Formato detectado: {formato_detectado["nome"]}', 'info')
            
            novos_colaboradores = []
            erros = []
            
            # Processa cada linha (pula o cabeçalho)
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row or len(row) < formato_detectado["min_colunas"]:
                    continue
                    
                # Extrai dados das colunas corretas baseado no formato detectado
                matricula = row[formato_detectado["col_matricula"]]  
                nome = row[formato_detectado["col_nome"]]       
                setor = row[formato_detectado["col_setor"]]
                
                # Valida dados obrigatórios
                if not matricula or not nome or not setor:
                    erros.append(f"Linha {row_num}: dados incompletos")
                    continue
                
                # Converte para string
                matricula = str(matricula).strip()
                nome = str(nome).strip()
                setor = str(setor).strip()
                
                # Verifica duplicatas na própria planilha
                if any(c['matricula'] == matricula for c in novos_colaboradores):
                    erros.append(f"Linha {row_num}: matrícula {matricula} duplicada na planilha")
                    continue
                
                novos_colaboradores.append({
                    'matricula': matricula,
                    'nome': nome,
                    'setor': setor
                })
            
            workbook.close()
            
            if erros:
                for erro in erros[:5]:  # Mostra apenas os primeiros 5 erros
                    flash(erro, 'warning')
                if len(erros) > 5:
                    flash(f'... e mais {len(erros) - 5} erros.', 'warning')
                return render_template('manager/upload.html', form=form, loja=current_user.loja)
            
            if not novos_colaboradores:
                flash('Nenhum colaborador válido encontrado na planilha.', 'warning')
                return render_template('manager/upload.html', form=form, loja=current_user.loja)
            
            # SUBSTITUI TODOS OS COLABORADORES
            # Remove todos os colaboradores existentes da loja (que não têm sorteios)
            colaboradores_existentes = Colaborador.query.filter_by(loja_id=current_user.loja_id).all()
            removidos = 0
            nao_removidos = 0
            colaboradores_protegidos = []
            
            for colaborador in colaboradores_existentes:
                tem_sorteios = SorteioColaborador.query.filter_by(colaborador_id=colaborador.id).first()
                if not tem_sorteios:
                    db.session.delete(colaborador)
                    removidos += 1
                else:
                    nao_removidos += 1
                    colaboradores_protegidos.append(colaborador.matricula)
            
            # Flush para garantir que as exclusões sejam processadas
            db.session.flush()
            
            # Adiciona novos colaboradores (verifica se não conflita com protegidos)
            adicionados = 0
            conflitos = []
            
            for dados in novos_colaboradores:
                # Verifica se matrícula já existe (colaborador protegido)
                if dados['matricula'] in colaboradores_protegidos:
                    conflitos.append(f"Matrícula {dados['matricula']} já existe (colaborador protegido)")
                    continue
                
                novo_colaborador = Colaborador(
                    matricula=dados['matricula'],
                    nome=dados['nome'],
                    setor=dados['setor'],
                    loja_id=current_user.loja_id,
                    apto=True,
                    ultima_atualizacao=get_brazil_datetime()
                )
                db.session.add(novo_colaborador)
                adicionados += 1
            
            db.session.commit()
            
            # Monta mensagem de sucesso
            mensagem = f"Upload concluído! {adicionados} colaboradores adicionados."
            if removidos > 0:
                mensagem += f" {removidos} colaboradores anteriores foram removidos."
            if nao_removidos > 0:
                mensagem += f" {nao_removidos} colaboradores com histórico foram mantidos."
            
            flash(mensagem, 'success')
            
            # Mostra conflitos se houver
            if conflitos:
                for conflito in conflitos[:3]:  # Mostra apenas os primeiros 3
                    flash(conflito, 'warning')
                if len(conflitos) > 3:
                    flash(f'... e mais {len(conflitos) - 3} conflitos de matrícula.', 'warning')
            
            return redirect(url_for('manager.colaboradores'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao processar arquivo: {str(e)}', 'danger')
    
    return render_template('manager/upload.html', form=form, loja=current_user.loja)

@manager_bp.route('/sortear', methods=['GET', 'POST'])
@manager_required
def sortear_colaboradores():
    """Sorteia 1 colaborador para o prêmio"""
    if not current_user.loja:
        flash('Você não está associado a nenhuma loja.', 'warning')
        return redirect(url_for('main.index'))
    
    # Verifica se a loja foi sorteada
    sorteio_atual = SorteioSemanal.query.order_by(SorteioSemanal.semana_inicio.desc()).first()
    
    if not sorteio_atual or (current_user.loja_id != sorteio_atual.loja_big_id and 
                           current_user.loja_id != sorteio_atual.loja_ultra_id):
        flash('Sua loja não foi sorteada esta semana.', 'info')
        return redirect(url_for('manager.dashboard'))
    
    form = SorteioColaboradorForm()
    
    # Popula prêmios disponíveis (ativos, futuros e específicos da loja ou gerais)
    # E que ainda não foram sorteados
    premios_ja_sorteados = db.session.query(SorteioColaborador.premio_id).join(Colaborador).filter(
        SorteioColaborador.sorteio_semanal_id == sorteio_atual.id,
        Colaborador.loja_id == current_user.loja_id
    ).subquery()
    
    premios_disponiveis = Premio.query.filter(
        Premio.ativo == True,
        Premio.data_evento >= date.today(),
        db.or_(Premio.loja_id == current_user.loja_id, Premio.loja_id.is_(None)),
        ~Premio.id.in_(db.select(premios_ja_sorteados).scalar_subquery())  # Exclui prêmios já sorteados
    ).order_by(Premio.data_evento).all()
    
    form.premio_id.choices = [(p.id, f"{p.nome} - {p.data_evento.strftime('%d/%m/%Y')}") for p in premios_disponiveis]
    
    if not premios_disponiveis:
        flash('Não há mais prêmios disponíveis para sorteio. Contate o administrador.', 'warning')
        return redirect(url_for('manager.dashboard'))
    
    if form.validate_on_submit():
        # Verifica novamente se já houve sorteio para este prêmio (dupla verificação)
        premio_selecionado = Premio.query.get(form.premio_id.data)
        ja_sorteado = SorteioColaborador.query.join(Colaborador).filter(
            SorteioColaborador.sorteio_semanal_id == sorteio_atual.id,
            SorteioColaborador.premio_id == form.premio_id.data,
            Colaborador.loja_id == current_user.loja_id
        ).first()
        
        if ja_sorteado:
            flash(f'Já foi realizado sorteio para o prêmio "{premio_selecionado.nome}" esta semana.', 'warning')
            return redirect(url_for('manager.dashboard'))
        
        # Busca colaboradores já sorteados nesta semana (para excluir da lista)
        colaboradores_ja_sorteados = db.session.query(SorteioColaborador.colaborador_id).join(Colaborador).filter(
            SorteioColaborador.sorteio_semanal_id == sorteio_atual.id,
            Colaborador.loja_id == current_user.loja_id
        ).subquery()
        
        # Busca colaboradores aptos (excluindo os já sorteados)
        colaboradores_aptos = Colaborador.query.filter_by(
            loja_id=current_user.loja_id,
            apto=True
        ).filter(
            ~Colaborador.id.in_(db.select(colaboradores_ja_sorteados).scalar_subquery())  # Exclui já sorteados
        ).all()
        
        if len(colaboradores_aptos) < 1:
            flash('Não há colaboradores disponíveis para sorteio (todos já foram sorteados).', 'warning')
            
            # Calcula informações de prêmios para a loja
            total_premios_loja = Premio.query.filter(
                Premio.ativo == True,
                Premio.data_evento >= date.today(),
                db.or_(Premio.loja_id == current_user.loja_id, Premio.loja_id.is_(None))
            ).count()
            
            premios_sorteados_count = db.session.query(SorteioColaborador).join(Colaborador).filter(
                SorteioColaborador.sorteio_semanal_id == sorteio_atual.id,
                Colaborador.loja_id == current_user.loja_id
            ).count()
            
            return render_template('manager/sortear.html', form=form,
                                 colaboradores_count=len(colaboradores_aptos),
                                 colaboradores=[],
                                 total_premios=total_premios_loja,
                                 premios_sorteados=premios_sorteados_count)
        
        # Cria snapshot da lista de colaboradores
        colaboradores_snapshot = [
            {
                'id': c.id,
                'matricula': c.matricula,
                'nome': c.nome,
                'setor': c.setor
            } for c in colaboradores_aptos
        ]
        
        # Verifica se veio do sorteio animado
        sorteio_animado = request.form.get('sorteio_animado') == 'true'
        colaborador_sorteado = None
        
        if sorteio_animado:
            # Pega o colaborador sorteado pela animação
            colaborador_id = request.form.get('colaborador_sorteado_id')
            if colaborador_id:
                colaborador_sorteado = Colaborador.query.filter_by(
                    id=int(colaborador_id),
                    loja_id=current_user.loja_id,
                    apto=True
                ).first()
                
                # Verifica se o colaborador não foi sorteado já
                if colaborador_sorteado and colaborador_sorteado.id in [sc.colaborador_id for sc in SorteioColaborador.query.join(Colaborador).filter(
                    SorteioColaborador.sorteio_semanal_id == sorteio_atual.id,
                    Colaborador.loja_id == current_user.loja_id
                ).all()]:
                    flash('O colaborador sorteado já foi contemplado esta semana.', 'warning')
                    return redirect(url_for('manager.dashboard'))
        
        if not colaborador_sorteado:
            # Realiza sorteio aleatório (sorteio simples)
            colaborador_sorteado = random.choice(colaboradores_aptos)
        
        # Registra o sorteado
        sorteio = SorteioColaborador(
            sorteio_semanal_id=sorteio_atual.id,
            premio_id=form.premio_id.data,
            colaborador_id=colaborador_sorteado.id,
            sorteado_por=current_user.id,
            lista_confirmada=True,  # Assistente confirmou
            colaboradores_snapshot=json.dumps(colaboradores_snapshot)
        )
        db.session.add(sorteio)
        db.session.commit()
        
        flash(f'🎉 Colaborador {colaborador_sorteado.nome} foi sorteado para "{premio_selecionado.nome}"!', 'success')
        return redirect(url_for('manager.dashboard'))
    
    # Busca colaboradores já sorteados nesta semana (para excluir da lista)
    colaboradores_ja_sorteados = db.session.query(SorteioColaborador.colaborador_id).join(Colaborador).filter(
        SorteioColaborador.sorteio_semanal_id == sorteio_atual.id,
        Colaborador.loja_id == current_user.loja_id
    ).subquery()
    
    # Busca colaboradores aptos para o frontend (excluindo os já sorteados)
    colaboradores_aptos = Colaborador.query.filter_by(
        loja_id=current_user.loja_id,
        apto=True
    ).filter(
        ~Colaborador.id.in_(colaboradores_ja_sorteados)  # Exclui já sorteados
    ).all()
    
    # Serializa colaboradores para JSON
    colaboradores_json = [
        {
            'id': c.id,
            'matricula': c.matricula,
            'nome': c.nome,
            'setor': c.setor
        } for c in colaboradores_aptos
    ]

    # Calcula informações de prêmios para a loja
    total_premios_loja = Premio.query.filter(
        Premio.ativo == True,
        Premio.data_evento >= date.today(),
        db.or_(Premio.loja_id == current_user.loja_id, Premio.loja_id.is_(None))
    ).count()
    
    premios_sorteados_count = db.session.query(SorteioColaborador).join(Colaborador).filter(
        SorteioColaborador.sorteio_semanal_id == sorteio_atual.id,
        Colaborador.loja_id == current_user.loja_id
    ).count()

    return render_template('manager/sortear.html',
                         form=form,
                         colaboradores_count=len(colaboradores_aptos),
                         colaboradores=colaboradores_json,
                         total_premios=total_premios_loja,
                         premios_sorteados=premios_sorteados_count)

@manager_bp.route('/sortear/ajax', methods=['POST'])
@manager_required
def sortear_colaboradores_ajax():
    """Sorteia colaborador via AJAX sem redirecionamento"""
    if not current_user.loja:
        return jsonify({'success': False, 'message': 'Você não está associado a nenhuma loja.'}), 400
    
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'success': False, 'message': 'Dados não fornecidos'}), 400
        
        premio_id = data.get('premio_id')
        colaborador_id = data.get('colaborador_id')
        
        if not all([premio_id, colaborador_id]):
            return jsonify({'success': False, 'message': 'Dados incompletos'}), 400
        
        # Verifica se a loja foi sorteada
        sorteio_atual = SorteioSemanal.query.order_by(SorteioSemanal.semana_inicio.desc()).first()
        
        if not sorteio_atual or (current_user.loja_id != sorteio_atual.loja_big_id and 
                               current_user.loja_id != sorteio_atual.loja_ultra_id):
            return jsonify({'success': False, 'message': 'Sua loja não foi sorteada esta semana.'}), 400
        
        # Busca prêmio
        premio = Premio.query.get(premio_id)
        if not premio:
            return jsonify({'success': False, 'message': 'Prêmio não encontrado'}), 400
        
        # Busca colaborador
        colaborador = Colaborador.query.filter_by(
            id=colaborador_id,
            loja_id=current_user.loja_id,
            apto=True
        ).first()
        
        if not colaborador:
            return jsonify({'success': False, 'message': 'Colaborador não encontrado ou não apto'}), 400
        
        # Verifica se já foi sorteado nesta semana
        ja_sorteado = SorteioColaborador.query.join(Colaborador).filter(
            SorteioColaborador.sorteio_semanal_id == sorteio_atual.id,
            SorteioColaborador.colaborador_id == colaborador.id,
            Colaborador.loja_id == current_user.loja_id
        ).first()
        
        if ja_sorteado:
            return jsonify({'success': False, 'message': 'Este colaborador já foi sorteado esta semana'}), 400
        
        # Verifica se já houve sorteio para este prêmio
        premio_ja_sorteado = SorteioColaborador.query.join(Colaborador).filter(
            SorteioColaborador.sorteio_semanal_id == sorteio_atual.id,
            SorteioColaborador.premio_id == premio.id,
            Colaborador.loja_id == current_user.loja_id
        ).first()
        
        if premio_ja_sorteado:
            return jsonify({'success': False, 'message': f'Já foi realizado sorteio para o prêmio "{premio.nome}" esta semana'}), 400
        
        # Busca todos os colaboradores aptos para snapshot
        colaboradores_ja_sorteados = db.session.query(SorteioColaborador.colaborador_id).join(Colaborador).filter(
            SorteioColaborador.sorteio_semanal_id == sorteio_atual.id,
            Colaborador.loja_id == current_user.loja_id
        ).subquery()
        
        colaboradores_aptos = Colaborador.query.filter_by(
            loja_id=current_user.loja_id,
            apto=True
        ).filter(
            ~Colaborador.id.in_(colaboradores_ja_sorteados)
        ).all()
        
        # Cria snapshot da lista
        colaboradores_snapshot = [
            {
                'id': c.id,
                'matricula': c.matricula,
                'nome': c.nome,
                'setor': c.setor
            } for c in colaboradores_aptos
        ]
        
        # Registra o sorteio
        sorteio = SorteioColaborador(
            sorteio_semanal_id=sorteio_atual.id,
            premio_id=premio.id,
            colaborador_id=colaborador.id,
            sorteado_por=current_user.id,
            lista_confirmada=True,
            colaboradores_snapshot=json.dumps(colaboradores_snapshot)
        )
        
        db.session.add(sorteio)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'🎉 Colaborador {colaborador.nome} foi sorteado para "{premio.nome}"!',
            'data': {
                'colaborador': {'nome': colaborador.nome, 'setor': colaborador.setor},
                'premio': {
                    'nome': premio.nome, 
                    'data_evento': premio.data_evento.strftime('%d/%m/%Y'),
                    'imagem_url': premio.get_imagem_url()
                }
            }
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Erro interno: {str(e)}'}), 500 